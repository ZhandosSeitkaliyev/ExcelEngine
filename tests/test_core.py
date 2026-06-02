import io
import json
import openpyxl
import pytest

from engine.core import (
    analyze,
    fill,
    schema_to_dict,
    wb_from_bytes,
    wb_to_bytes,
    MARKER_ARGB,
)


# ──────────────────────────────────────────
# analyze — schema extraction
# ──────────────────────────────────────────

def test_analyze_fillable_fields(template_wb):
    schema = analyze(template_wb)
    names = [f.name for f in schema.fillable_fields]

    assert "Name" in names
    assert "Amount" in names
    # Status has a marker fill on the *header*, but the column is still fillable
    assert "Status" in names
    # Service column A has no text header → must NOT appear in fillable_fields
    assert all(f.col_letter != "A" for f in schema.fillable_fields)


def test_analyze_service_cols(template_wb):
    schema = analyze(template_wb)
    # Column A (index 1) has an empty header → service column
    assert 1 in schema.service_cols
    # Columns with text headers are NOT service columns
    assert 2 not in schema.service_cols  # B — Name
    assert 3 not in schema.service_cols  # C — Amount
    assert 4 not in schema.service_cols  # D — Status


def test_analyze_protected_cells(template_wb):
    schema = analyze(template_wb)
    coords = [p.coordinate for p in schema.protected_cells]

    assert "A1" in coords   # service col header
    assert "D1" in coords   # fillable header with marker
    assert "A2" in coords   # merged structural cell
    assert "C3" in coords   # data-area pre-filled cell


def test_analyze_no_marker_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"].value = "Header"
    with pytest.raises(ValueError, match="не размечен маркером"):
        analyze(wb)


# ──────────────────────────────────────────
# fill — writing data
# ──────────────────────────────────────────

def test_fill_writes_to_fillable_cells(template_bytes):
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)

    data = [{"Name": "Alice", "Amount": 1500.0, "Status": "Active"}]
    wb = fill(wb, schema, data)
    ws = wb.active

    name_col   = next(f.col_index for f in schema.fillable_fields if f.name == "Name")
    amount_col = next(f.col_index for f in schema.fillable_fields if f.name == "Amount")
    status_col = next(f.col_index for f in schema.fillable_fields if f.name == "Status")
    data_row = schema.header_row + 1

    assert ws.cell(row=data_row, column=name_col).value == "Alice"
    assert ws.cell(row=data_row, column=amount_col).value == 1500.0
    assert ws.cell(row=data_row, column=status_col).value == "Active"


def test_fill_never_touches_header_row(template_bytes):
    """Header row values must be identical before and after fill(), for all columns."""
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)

    ws = wb.active
    header_before = {cell.coordinate: cell.value for cell in ws[schema.header_row]}

    data = [{"Name": "Bob", "Amount": 999, "Status": "Shipped"}]
    wb = fill(wb, schema, data)
    ws = wb.active

    for coord, original in header_before.items():
        assert ws[coord].value == original, f"Header cell {coord} was modified!"


def test_fill_skips_protected_data_cells(template_bytes):
    """
    C3 carries a marker and the value 'FORMULA'.
    When filling two records (rows 2 and 3), row-3 col-C must retain 'FORMULA'.
    """
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)

    data = [{"Name": "Row2", "Amount": 100.0}, {"Name": "Row3", "Amount": 999.0}]
    wb = fill(wb, schema, data)
    ws = wb.active

    amount_col = next(f.col_index for f in schema.fillable_fields if f.name == "Amount")
    assert ws.cell(row=3, column=amount_col).value == "FORMULA"
    # Row 2 col C is unprotected → should be written
    assert ws.cell(row=2, column=amount_col).value == 100.0


def test_fill_removes_marker_fill(template_bytes):
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)
    data = [{"Name": "Carol", "Amount": 42.0, "Status": "Done"}]
    result_wb = wb_from_bytes(wb_to_bytes(fill(wb, schema, data)))
    ws = result_wb.active

    for row in ws.iter_rows():
        for cell in row:
            if cell.fill and cell.fill.fill_type == "solid":
                assert cell.fill.fgColor.rgb.upper() != MARKER_ARGB, (
                    f"Marker fill still present at {cell.coordinate}"
                )


def test_fill_preserves_non_fill_formatting(template_bytes):
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)
    result_wb = wb_from_bytes(wb_to_bytes(fill(wb, schema, [{"Name": "Dave", "Amount": 7777.77}])))
    ws = result_wb.active

    b1 = ws["B1"]
    assert b1.font.bold is True,              "Bold font on B1 was lost"
    assert b1.border.left.border_style == "thin", "Border on B1 was lost"
    assert b1.alignment.horizontal == "center",   "Alignment on B1 was lost"

    c1 = ws["C1"]
    assert c1.number_format == "#,##0.00",    "Number format on C1 was lost"


def test_fill_multiple_records(template_bytes):
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)

    data = [
        {"Name": "Row1", "Amount": 1.0},
        {"Name": "Row2", "Amount": 2.0},
        {"Name": "Row3", "Amount": 3.0},
    ]
    wb = fill(wb, schema, data)
    ws = wb.active

    name_col = next(f.col_index for f in schema.fillable_fields if f.name == "Name")
    start = schema.header_row + 1

    assert ws.cell(row=start,     column=name_col).value == "Row1"
    assert ws.cell(row=start + 1, column=name_col).value == "Row2"
    assert ws.cell(row=start + 2, column=name_col).value == "Row3"


def test_fill_ignores_unknown_fields(template_bytes):
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)

    data = [{"Name": "Eve", "Amount": 5.0, "NonExistent": "IGNORED"}]
    wb = fill(wb, schema, data)
    ws = wb.active

    used_cols = {cell.column for row in ws.iter_rows() for cell in row if cell.value is not None}
    assert max(used_cols) <= 4, "A new column was created for an unknown field"


def test_fill_auto_number(template_bytes):
    """
    Service column A receives 1-based sequential numbers.
    A2 is protected (marker) → skipped.
    A3 is a non-top-left cell of the merged range A2:A3 → skipped.
    A4 and A5 (rows offset 2 and 3) are unprotected → get values 3 and 4.
    """
    wb = wb_from_bytes(template_bytes)
    schema = analyze(wb)

    data = [{"Name": "R1"}, {"Name": "R2"}, {"Name": "R3"}, {"Name": "R4"}]
    wb = fill(wb, schema, data, auto_number=True)
    ws = wb.active

    assert ws.cell(row=4, column=1).value == 3
    assert ws.cell(row=5, column=1).value == 4


# ──────────────────────────────────────────
# Serialisation
# ──────────────────────────────────────────

def test_schema_to_dict(template_wb):
    schema = analyze(template_wb)
    d = schema_to_dict(schema)
    serialized = json.dumps(d)
    parsed = json.loads(serialized)

    assert parsed["sheet"] == "Sheet1"
    assert any(f["name"] == "Name"   for f in parsed["fillable_fields"])
    assert any(f["name"] == "Status" for f in parsed["fillable_fields"])
    assert 1 in parsed["service_cols"]
