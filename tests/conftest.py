"""
Fixture sheet layout (1-indexed):

  Row 1 — headers
    A1: empty  + MARKER_FILL   → service column (auto-number, no field name)
    B1: "Name" + bold/border/center  → fillable
    C1: "Amount" + bold/border + #,##0.00  → fillable
    D1: "Status" + MARKER_FILL  → fillable (marker on header does NOT exclude column!)

  Data zone
    A2:A3 merged, value "§", MARKER_FILL  → structural protected cell
    C3: value "FORMULA", MARKER_FILL  → pre-filled/formula data cell (fill must skip it)

Expected analyze() result:
  fillable_fields : Name(B), Amount(C), Status(D)
  service_cols    : [1]  (column A)
  protected_cells : A1, D1, A2, C3
"""

import io
import pytest
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

MARKER_FILL = PatternFill(fill_type="solid", fgColor="FF00A0F0")
THIN = Side(border_style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BOLD = Font(bold=True)


def _make_template() -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # A1: service column — empty header, marked
    ws["A1"].fill = MARKER_FILL

    # B1: fillable with formatting to verify preservation
    ws["B1"].value = "Name"
    ws["B1"].font = BOLD
    ws["B1"].border = BORDER
    ws["B1"].alignment = Alignment(horizontal="center")

    # C1: fillable with number format
    ws["C1"].value = "Amount"
    ws["C1"].font = BOLD
    ws["C1"].border = BORDER
    ws["C1"].number_format = "#,##0.00"

    # D1: fillable header that also carries a marker (protection of header value,
    #     but the *column* is still fillable in data rows)
    ws["D1"].value = "Status"
    ws["D1"].fill = MARKER_FILL

    # A2:A3 merged structural cell — protected
    ws.merge_cells("A2:A3")
    ws["A2"].value = "§"
    ws["A2"].fill = MARKER_FILL

    # C3: pre-filled data cell — protected (simulate a formula placeholder)
    ws["C3"].value = "FORMULA"
    ws["C3"].fill = MARKER_FILL

    return wb


@pytest.fixture
def template_bytes() -> bytes:
    wb = _make_template()
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


@pytest.fixture
def template_wb() -> openpyxl.Workbook:
    return _make_template()
