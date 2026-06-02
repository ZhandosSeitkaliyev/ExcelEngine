from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

MARKER_ARGB = "FF00A0F0"
NO_FILL = PatternFill(fill_type=None)


# ──────────────────────────────────────────
# Domain types
# ──────────────────────────────────────────

@dataclass
class FillableField:
    name: str
    col_letter: str
    col_index: int


@dataclass
class ProtectedCell:
    coordinate: str
    col_letter: str
    row: int


@dataclass
class TemplateSchema:
    sheet: str
    header_row: int
    fillable_fields: list[FillableField]
    protected_cells: list[ProtectedCell]
    # Columns whose header cell is empty — not exposed as fillable fields.
    # Can receive auto-incremented numbers when fill() is called with auto_number=True.
    service_cols: list[int] = field(default_factory=list)


# ──────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────

def _is_marker(cell) -> bool:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return False
    fg = fill.fgColor
    if fg is None:
        return False
    return fg.type == "rgb" and fg.rgb.upper() == MARKER_ARGB


def _is_nontopleft_merged(ws, row: int, col: int) -> bool:
    """Return True if (row, col) is inside a merged range but is NOT the top-left cell."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            if not (row == rng.min_row and col == rng.min_col):
                return True
    return False


def _first_data_row(ws, schema: TemplateSchema) -> int:
    """Return the first row after the header block where all fillable columns are empty."""
    row = schema.header_row + 1
    while row <= ws.max_row:
        has_data = any(
            ws.cell(row=row, column=f.col_index).value is not None
            for f in schema.fillable_fields
        )
        if not has_data:
            break
        row += 1
    return row


# ──────────────────────────────────────────
# Public API
# ──────────────────────────────────────────

def analyze(wb: Workbook, header_row: int = 1) -> TemplateSchema:
    """
    Scan the first sheet of wb and return a TemplateSchema.

    fillable_fields  — every column in header_row that has a non-empty text value,
                       regardless of whether that cell is marked with the protection colour.
    protected_cells  — every cell anywhere in the sheet that carries the #00A0F0 fill.
                       During fill() these coordinates are never overwritten.
    service_cols     — columns whose header cell is empty/blank; excluded from the
                       fillable schema but can receive auto-numbering.

    Raises ValueError if no marker cells are found (template not annotated).
    """
    ws = wb.active
    protected: list[ProtectedCell] = []
    fillable: list[FillableField] = []
    service_cols: list[int] = []

    for row in ws.iter_rows():
        for cell in row:
            if _is_marker(cell):
                protected.append(
                    ProtectedCell(
                        coordinate=cell.coordinate,
                        col_letter=get_column_letter(cell.column),
                        row=cell.row,
                    )
                )

    if not protected:
        raise ValueError(
            "Шаблон не размечен маркером защиты: ни одна ячейка не имеет заливки #00A0F0."
        )

    for cell in ws[header_row]:
        col_letter = get_column_letter(cell.column)
        text = str(cell.value).strip() if cell.value is not None else ""
        if not text:
            service_cols.append(cell.column)
        else:
            fillable.append(
                FillableField(
                    name=text,
                    col_letter=col_letter,
                    col_index=cell.column,
                )
            )

    return TemplateSchema(
        sheet=ws.title,
        header_row=header_row,
        fillable_fields=fillable,
        protected_cells=protected,
        service_cols=service_cols,
    )


def fill(
    wb: Workbook,
    schema: TemplateSchema,
    data: list[dict[str, Any]],
    auto_number: bool = False,
) -> Workbook:
    """
    Write data rows into wb starting from the first free row after the header.

    Rules:
    - The header row is never touched.
    - A cell is skipped if its coordinate appears in protected_cells (marker fill).
    - Non-top-left cells of merged ranges are always skipped.
    - Fields in data that have no matching column are ignored silently.
    - If auto_number=True, service_cols receive 1-based sequential integers
      (unless the specific cell is protected or non-top-left merged).
    - After writing, the marker fill is removed from every protected cell.
    """
    ws = wb.active
    protected_coords = {p.coordinate for p in schema.protected_cells}
    col_map = {f.name: f.col_index for f in schema.fillable_fields}

    start_row = _first_data_row(ws, schema)

    for offset, record in enumerate(data):
        row_idx = start_row + offset

        if auto_number:
            for col_idx in schema.service_cols:
                if _is_nontopleft_merged(ws, row_idx, col_idx):
                    continue
                coord = f"{get_column_letter(col_idx)}{row_idx}"
                if coord not in protected_coords:
                    ws.cell(row=row_idx, column=col_idx, value=offset + 1)

        for field_name, value in record.items():
            col_idx = col_map.get(field_name)
            if col_idx is None:
                continue
            if _is_nontopleft_merged(ws, row_idx, col_idx):
                continue
            coord = f"{get_column_letter(col_idx)}{row_idx}"
            if coord in protected_coords:
                continue
            ws.cell(row=row_idx, column=col_idx, value=value)

    for pc in schema.protected_cells:
        ws[pc.coordinate].fill = NO_FILL

    return wb


# ──────────────────────────────────────────
# Serialisation helpers
# ──────────────────────────────────────────

def schema_to_dict(schema: TemplateSchema) -> dict:
    return {
        "sheet": schema.sheet,
        "header_row": schema.header_row,
        "fillable_fields": [
            {"name": f.name, "col_letter": f.col_letter, "col_index": f.col_index}
            for f in schema.fillable_fields
        ],
        "protected_cells": [
            {"coordinate": p.coordinate, "col_letter": p.col_letter, "row": p.row}
            for p in schema.protected_cells
        ],
        "service_cols": schema.service_cols,
    }


def wb_to_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def wb_from_bytes(data: bytes) -> Workbook:
    return openpyxl.load_workbook(io.BytesIO(data))
